# -*- coding: utf-8 -*-
"""
Spyder Editor

Author: Krishna Moorthy D.
"""

import pandas as pd
import numpy as np
import openpyxl
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import tkinter as tk

## Opening Webpage for Data Analysis
browser = webdriver.Chrome("C:/Users/kdharmalingam/test/chromedriver.exe")
browser.get('https://www.google.co.in')

## Opening Dataframe for Input Values
def Progress():
    file_pth = Samp2.get()
    test_file = pd.read_excel(file_pth)
    test_file = test_file.replace(np.nan,'')
    print(test_file)
    Timezone = test_file['Time Zone'].values
    Tim = test_file['Time'].values
    max_row = Timezone.size
    print(max_row)
    
    ## Getting Time Zone Wise Time
    for ROW in range(max_row):
        
        Time_Zone = Timezone[ROW]
        ele1 = browser.find_element_by_name('q')
        ele1.clear()
        ele1.send_keys((Time_Zone + ' Time Now') + Keys.RETURN)
        ele2 = browser.find_element_by_id('rso').find_elements_by_tag_name('DIV')[3]
        Tval = ele2.get_attribute("innerText")
        print(str(Tval))
        Tval = datetime.strptime(Tval, "%I:%M %p")
        Tval = Tval.strftime("%H:%M:%S")
        Tim[ROW] = Tval
    
    print(test_file)
    
    browser.quit()
    
    ## Transfer Frame data to Excel Workbook
    wb = openpyxl.Workbook()
    ws = wb.sheetnames
    print(ws)
    
    TimeZ = test_file['Time Zone'].values
    Tme = test_file['Time'].values
    
    for ROWW in range(max_row):
        
        wb['Sheet'].cell(ROWW + 1, 1).value = TimeZ[ROWW]
        wb['Sheet'].cell(ROWW + 1, 2).value = Tme[ROWW]
    
    res_name = Samp4.get()
    wb.save('C:/Users/kdharmalingam/test/' + res_name + '.xlsx')
    wdow.quit()

## Main GUI For Process Inputs
wdow = tk.Tk()
wdow.title("Process Input")
Samp = tk.Frame(wdow)
Samp.pack()
Samp1 = tk.Label(Samp, text="File Path")
Samp2 = tk.Entry(Samp, width=100)
Samp3 = tk.Label(Samp, text="Result File Name")
Samp4 = tk.Entry(Samp, width=50)
Samp5 = tk.Button(Samp, text="Click", command=Progress)

Samp1.grid(column=0,row=0)
Samp2.grid(column=1,row=0)
Samp3.grid(column=0,row=2)
Samp4.grid(column=1,row=2)
Samp5.grid(column=1,row=5)

wdow.mainloop()
